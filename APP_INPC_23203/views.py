from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
import json
from django.db.models import Sum, Avg
from datetime import datetime, date, timedelta  # Added import for date and timedelta
import random  # Added import for random

# Import de tous vos modèles
from .models import (
    ProductType, Product, Wilaya, Moughataa, Commune, PointOfSale,
    ProductPrice, Cart, CartProducts, Sale, INPC  # <-- Sale et INPC uniquement si vous les avez
)

# Import de tous vos formulaires
from .forms import (
    CartForm, ProductForm, ProductTypeForm, WilayaForm,
    MoughataaForm, CommuneForm, ProductPriceForm,
    PointOfSaleForm, CartProductsForm
)

# =========================== #
#   PAGES PRINCIPALES         #
# =========================== #

def calculate_inpc(cart, start_date, end_date):
    total_weighted_price_change = 0.0
    total_weight = 0.0

    # Convertir start_date et end_date en datetime si ce sont des objets de type date
    if isinstance(start_date, date):
        start_date = datetime.combine(start_date, datetime.min.time())
    if isinstance(end_date, date):
        end_date = datetime.combine(end_date, datetime.min.time())

    # Parcourir tous les produits dans le panier
    for cart_product in cart.cartproducts_set.all():
        product = cart_product.product
        weight = float(cart_product.weight)  # Convertir le poids en float
        date_from = cart_product.date_from
        date_to = cart_product.date_to

        # Convertir date_from et date_to en datetime si ce sont des objets de type date
        if isinstance(date_from, date):
            date_from = datetime.combine(date_from, datetime.min.time())
        if isinstance(date_to, date):
            date_to = datetime.combine(date_to, datetime.min.time())

        # Vérifier que le produit est dans la période demandée
        if (date_from <= end_date) and (date_to >= start_date):
            # Récupérer les prix du produit pour la période donnée
            prices = ProductPrice.objects.filter(
                product=product,
                date_from__lte=end_date,
                date_to__gte=start_date
            ).order_by('date_from')

            if prices.exists():
                # Récupérer le premier prix (initial) et le dernier prix (final)
                initial_price = prices.first().value
                final_price = prices.last().value

                # Assurez-vous que les prix sont valides avant de calculer la variation
                if initial_price is not None and final_price is not None and initial_price != 0:
                    price_change = ((final_price - initial_price) / initial_price) * 100

                    # Ajouter à la somme pondérée
                    total_weighted_price_change += price_change * weight
                    total_weight += weight

    # Calculer l'INPC
    if total_weight > 0:
        inpc_value = total_weighted_price_change / total_weight
    else:
        inpc_value = 0

    return inpc_value

def home(request):
    """
    Vue qui affiche l'INPC pour un panier sélectionné par code, mois et année.
    """
    # Obtenir le panier sélectionné par code ou le premier panier par défaut
    cart_code = request.GET.get('cart_code')
    month = request.GET.get('month')
    year = request.GET.get('year')

    if cart_code:
        cart = Cart.objects.filter(code=cart_code).first()
    else:
        cart = Cart.objects.first()

    if month and year:
        month = int(month)  # Convert month to integer
        year = int(year)  # Convert year to integer
        start_date = datetime(year, month, 1)
    else:
        end_date = datetime.now()
        start_date = end_date.replace(day=1)
        month = start_date.month
        year = start_date.year

    end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    # Calcul de l'INPC pour le mois sélectionné
    inpc_value = calculate_inpc(cart, start_date, end_date)

    context = {
        'inpc_value': inpc_value,
        'carts': Cart.objects.all(),  # Pass all carts to the template
        'selected_cart': cart,  # Pass the selected cart to the template
        'selected_month': f"{month:02d}",  # Format month as two digits
        'selected_year': year,
        'months': [
            {'num': 1, 'name': 'Jan'},
            {'num': 2, 'name': 'Fév'},
            {'num': 3, 'name': 'Mar'},
            {'num': 4, 'name': 'Avr'},
            {'num': 5, 'name': 'Mai'},
            {'num': 6, 'name': 'Juin'},
            {'num': 7, 'name': 'Juil'},
            {'num': 8, 'name': 'Aoû'},
            {'num': 9, 'name': 'Sep'},
            {'num': 10, 'name': 'Oct'},
            {'num': 11, 'name': 'Nov'},
            {'num': 12, 'name': 'Déc'}
        ],  # Ensure months is a list of dictionaries with month numbers and names
        'years': list(range(2020, 2026))  # Ensure years is a list
    }

    # Generate random sales data for the line chart
    line_labels = ["Jan", "Fév", "Mar", "Avr", "Mai", "Juin", "Juil", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    line_data = [random.randint(500000, 2000000) for _ in line_labels]  # Values in Mauritanian Ouguiya (MRO)

    # Generate random product type data for the pie chart
    product_types = ["Alimentation", "Logement", "Transport", "Santé"]
    pie_labels = product_types
    pie_data = [random.randint(100000, 500000) for _ in product_types]  # Values in MRO

    def get_quarter(month):
        mapping = {
            'Jan': 'Q1', 'Fév': 'Q1', 'Mar': 'Q1',
            'Avr': 'Q2', 'Mai': 'Q2', 'Juin': 'Q2',
            'Juil': 'Q3', 'Aoû': 'Q3', 'Sep': 'Q3',
            'Oct': 'Q4', 'Nov': 'Q4', 'Déc': 'Q4',
        }
        return mapping.get(month, 'Autres')

    # Generate random quarterly data for the bar chart
    bar_dict = {}
    for month in line_labels:
        q = get_quarter(month)
        bar_dict[q] = bar_dict.get(q, 0) + random.randint(500000, 2000000)  # Values in MRO

    bar_labels = list(bar_dict.keys())
    bar_data = list(bar_dict.values())

    context.update({
        'line_labels': json.dumps(line_labels),
        'line_data': json.dumps(line_data),
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
        'bar_labels': json.dumps(bar_labels),
        'bar_data': json.dumps(bar_data),
    })
    return render(request, 'home.html', context)

def dashboard(request):
    return render(request, 'dashboard.html')

def contact(request):
    return render(request, 'contact.html')

# =========================== #
#   LISTES DES TABLES (READ)  #
# =========================== #

def cart_list(request):
    carts = Cart.objects.all()
    return render(request, "cart_list.html", {"cart_list": carts})

def product_type_list(request):
    product_types = ProductType.objects.all()
    return render(request, "producttype_list.html", {"product_types": product_types})

def product_list(request):
    products = Product.objects.all()
    return render(request, "product_list.html", {"product_list": products})

def wilaya_list(request):
    wilayas = Wilaya.objects.all()
    return render(request, "wilaya_list.html", {"wilaya_list": wilayas})

def moughataa_list(request):
    moughataas = Moughataa.objects.all()
    return render(request, "moughataa_list.html", {"moughataa_list": moughataas})

def commune_list(request):
    communes = Commune.objects.all()
    return render(request, "commune_list.html", {"commune_list": communes})

def cart_products_list(request):
    cart_products = CartProducts.objects.all()
    return render(request, "cartproducts_list.html", {"cartproducts_list": cart_products})

def point_of_sale_list(request):
    points_of_sale = PointOfSale.objects.all()
    return render(request, "pointofsale_list.html", {"pointofsale_list": points_of_sale})

def product_price_list(request):
    prices = ProductPrice.objects.all()
    return render(request, "productprice_list.html", {"productprice_list": prices})

def cartproduct_list(request):
    cartproducts = CartProducts.objects.all()
    context = {
        'cartproduct_list': cartproducts
    }
    return render(request, 'cartproduct_list.html', context)

# =========================== #
#   CRUD - CART (Panier)      #
# =========================== #

def cart_add(request):
    if request.method == "POST":
        form = CartForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Panier ajouté avec succès !")
            return redirect("cart_list")
    else:
        form = CartForm()
    return render(request, "cart_form.html", {"form": form})

def cart_edit(request, pk):
    cart = get_object_or_404(Cart, pk=pk)
    if request.method == "POST":
        form = CartForm(request.POST, instance=cart)
        if form.is_valid():
            form.save()
            messages.success(request, "Panier modifié avec succès !")
            return redirect("cart_list")
    else:
        form = CartForm(instance=cart)
    return render(request, "cart_form.html", {"form": form})

def cart_delete(request, pk):
    cart = get_object_or_404(Cart, pk=pk)
    if request.method == "POST":
        cart.delete()
        messages.success(request, "Panier supprimé avec succès !")
        return redirect("cart_list")
    return render(request, "cart_confirm_delete.html", {"cart": cart})

def cart_export(request):
    """
    Exporte la table Cart dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cart"

    # Ajouter l'en-tête
    fields = ["code", "name", "description"]
    ws.append(fields)

    # Ajouter les données
    for cart in Cart.objects.all():
        row_data = [cart.code, cart.name, cart.description]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cart_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - PRODUCT            #
# =========================== #

def product_add(request):
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit ajouté avec succès !")
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(request, "product_form.html", {"form": form})

def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit modifié avec succès !")
            return redirect("product_list")
    else:
        form = ProductForm(instance=product)
    return render(request, "product_form.html", {"form": form})

def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.delete()
        messages.success(request, "Produit supprimé avec succès !")
        return redirect("product_list")
    return render(request, "product_confirm_delete.html", {"product": product})

def product_export(request):
    """
    Exporte la table Product dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product"

    # Ajouter l'en-tête
    fields = ["code", "name", "description", "unit_measure", "product_type"]
    ws.append(fields)

    # Ajouter les données
    for product in Product.objects.all():
        row_data = [
            product.code,
            product.name,
            product.description,
            product.unit_measure,
            product.product_type.label  # Changed from product.product_type.name to product.product_type.label
        ]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="product_export.xlsx"'
    wb.save(response)
    return response

def product_import(request):
    """
    Importe la table Product depuis un fichier Excel
    (code, name, description, unit_measure, product_type).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                unit_val = row[3]
                product_type_label = row[4]

                product_type = ProductType.objects.get(label=product_type_label)

                Product.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val,
                    unit_measure=unit_val,
                    product_type=product_type
                )
            messages.success(request, "Importation des Produits réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("product_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("product_list")

# =========================== #
#   CRUD - PRODUCT TYPE       #
# =========================== #

def product_type_add(request):
    if request.method == "POST":
        form = ProductTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Type de produit ajouté avec succès !")
            return redirect("product_type_list")
    else:
        form = ProductTypeForm()
    return render(request, "producttype_form.html", {"form": form})

def product_type_edit(request, pk):
    product_type = get_object_or_404(ProductType, pk=pk)
    if request.method == "POST":
        form = ProductTypeForm(request.POST, instance=product_type)
        if form.is_valid():
            form.save()
            messages.success(request, "Type de produit modifié avec succès !")
            return redirect("product_type_list")
    else:
        form = ProductTypeForm(instance=product_type)
    return render(request, "producttype_form.html", {"form": form})

def product_type_delete(request, pk):
    product_type = get_object_or_404(ProductType, pk=pk)
    if request.method == "POST":
        product_type.delete()
        messages.success(request, "Type de produit supprimé avec succès !")
        return redirect("product_type_list")
    return render(request, "producttype_confirm_delete.html", {"product_type": product_type})

def product_type_export(request):
    """
    Exporte la table ProductType dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProductType"

    # Ajouter l'en-tête
    fields = ["code", "label", "description"]
    ws.append(fields)

    # Ajouter les données
    for product_type in ProductType.objects.all():
        row_data = [product_type.code, product_type.label, product_type.description]  # Changed from product_type.name to product_type.label
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="product_type_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - WILAYA             #
# =========================== #

def wilaya_add(request):
    if request.method == "POST":
        form = WilayaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Wilaya ajoutée avec succès !")
            return redirect("wilaya_list")
    else:
        form = WilayaForm()
    return render(request, "wilaya_form.html", {"form": form})

def wilaya_edit(request, pk):
    wilaya = get_object_or_404(Wilaya, pk=pk)
    if request.method == "POST":
        form = WilayaForm(request.POST, instance=wilaya)
        if form.is_valid():
            form.save()
            messages.success(request, "Wilaya modifiée avec succès !")
            return redirect("wilaya_list")
    else:
        form = WilayaForm(instance=wilaya)
    return render(request, "wilaya_form.html", {"form": form})

def wilaya_delete(request, pk):
    wilaya = get_object_or_404(Wilaya, pk=pk)
    if request.method == "POST":
        wilaya.delete()
        messages.success(request, "Wilaya supprimée avec succès !")
        return redirect("wilaya_list")
    return render(request, "wilaya_confirm_delete.html", {"wilaya": wilaya})

def wilaya_import(request):
    """
    Importe la table Wilaya depuis un fichier Excel
    (code, name).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                Wilaya.objects.create(
                    code=code_val,
                    name=name_val
                )
            messages.success(request, "Importation des Wilayas réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("wilaya_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("wilaya_list")

def wilaya_export(request):
    """
    Exporte la table Wilaya dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wilaya"

    # Ajouter l'en-tête
    fields = ["code", "name"]
    ws.append(fields)

    # Ajouter les données
    for wilaya in Wilaya.objects.all():
        row_data = [wilaya.code, wilaya.name]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="wilaya_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - MOUGHATAA          #
# =========================== #

def moughataa_add(request):
    if request.method == "POST":
        form = MoughataaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Moughataa ajoutée avec succès !")
            return redirect("moughataa_list")
    else:
        form = MoughataaForm()
    return render(request, "moughataa_form.html", {"form": form})

def moughataa_edit(request, pk):
    moughataa = get_object_or_404(Moughataa, pk=pk)
    if request.method == "POST":
        form = MoughataaForm(request.POST, instance=moughataa)
        if form.is_valid():
            form.save()
            messages.success(request, "Moughataa modifiée avec succès !")
            return redirect("moughataa_list")
    else:
        form = MoughataaForm(instance=moughataa)
    return render(request, "moughataa_form.html", {"form": form})

def moughataa_delete(request, pk):
    moughataa = get_object_or_404(Moughataa, pk=pk)
    if request.method == "POST":
        moughataa.delete()
        messages.success(request, "Moughataa supprimée avec succès !")
        return redirect("moughataa_list")
    return render(request, "moughataa_confirm_delete.html", {"moughataa": moughataa})

def moughataa_import(request):
    """
    Importe la table Moughataa depuis un fichier Excel
    (code, label, wilaya).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                label_val = row[1]
                wilaya_name = row[2]

                wilaya = Wilaya.objects.get(name=wilaya_name)

                Moughataa.objects.create(
                    code=code_val,
                    label=label_val,
                    wilaya=wilaya
                )
            messages.success(request, "Importation des Moughataas réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("moughataa_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("moughataa_list")

def moughataa_export(request):
    """
    Exporte la table Moughataa dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Moughataa"

    # Ajouter l'en-tête
    fields = ["code", "label", "wilaya"]
    ws.append(fields)

    # Ajouter les données
    for moughataa in Moughataa.objects.all():
        row_data = [moughataa.code, moughataa.label, moughataa.wilaya.name]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="moughataa_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - COMMUNE            #
# =========================== #

def commune_add(request):
    if request.method == "POST":
        form = CommuneForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Commune ajoutée avec succès !")
            return redirect("commune_list")
    else:
        form = CommuneForm()
    return render(request, "commune_form.html", {"form": form})

def commune_edit(request, pk):
    commune = get_object_or_404(Commune, pk=pk)
    if request.method == "POST":
        form = CommuneForm(request.POST, instance=commune)
        if form.is_valid():
            form.save()
            messages.success(request, "Commune modifiée avec succès !")
            return redirect("commune_list")
    else:
        form = CommuneForm(instance=commune)
    return render(request, "commune_form.html", {"form": form})

def commune_delete(request, pk):
    commune = get_object_or_404(Commune, pk=pk)
    if request.method == "POST":
        commune.delete()
        messages.success(request, "Commune supprimée avec succès !")
        return redirect("commune_list")
    return render(request, "commune_confirm_delete.html", {"commune": commune})

def commune_import(request):
    """
    Importe la table Commune depuis un fichier Excel
    (code, name, moughataa).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                moughataa_name = row[2]

                moughataa = Moughataa.objects.get(name=moughataa_name)

                Commune.objects.create(
                    code=code_val,
                    name=name_val,
                    moughataa=moughataa
                )
            messages.success(request, "Importation des Communes réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("commune_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("commune_list")

def commune_export(request):
    """
    Exporte la table Commune dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commune"

    # Ajouter l'en-tête
    fields = ["code", "name", "moughataa"]
    ws.append(fields)

    # Ajouter les données
    for commune in Commune.objects.all():
        row_data = [commune.code, commune.name, commune.moughataa.label]  # Changed from commune.moughataa.name to commune.moughataa.label
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="commune_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - POINT OF SALE      #
# =========================== #

def point_of_sale_add(request):
    if request.method == "POST":
        form = PointOfSaleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Point de vente créé avec succès.")
            return redirect("point_of_sale_list")
    else:
        form = PointOfSaleForm()
    return render(request, "pointofsale_form.html", {"form": form})

def point_of_sale_edit(request, pk):
    pos = get_object_or_404(PointOfSale, pk=pk)
    if request.method == "POST":
        form = PointOfSaleForm(request.POST, instance=pos)
        if form.is_valid():
            form.save()
            messages.success(request, "Point de vente modifié avec succès.")
            return redirect("point_of_sale_list")
    else:
        form = PointOfSaleForm(instance=pos)
    return render(request, "pointofsale_form.html", {"form": form})

def point_of_sale_delete(request, pk):
    pos = get_object_or_404(PointOfSale, pk=pk)
    if request.method == "POST":
        pos.delete()
        messages.success(request, "Point de vente supprimé avec succès.")
        return redirect("point_of_sale_list")
    return render(request, "pointofsale_confirm_delete.html", {"point_of_sale": pos})

def point_of_sale_import(request):
    """
    Importe la table PointOfSale depuis un fichier Excel
    (code, name, description).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                PointOfSale.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val
                )
            messages.success(request, "Importation des Points de Vente réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("point_of_sale_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("point_of_sale_list")

def point_of_sale_export(request):
    """
    Exporte la table PointOfSale dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PointOfSale"

    # Ajouter l'en-tête
    fields = ["code", "type", "gps_lat", "gps_lon", "commune"]
    ws.append(fields)

    # Ajouter les données
    for point in PointOfSale.objects.all():
        row_data = [point.code, point.type, point.gps_lat, point.gps_lon, point.commune.label]  # Changed from point.description to point.gps_lat, point.gps_lon, point.commune.label
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="point_of_sale_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - PRODUCT PRICE      #
# =========================== #

def product_price_add(request):
    if request.method == "POST":
        form = ProductPriceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Le prix a été créé avec succès.")
            return redirect("product_price_list")
    else:
        form = ProductPriceForm()
    return render(request, "productprice_form.html", {"form": form})

def product_price_edit(request, pk):
    price = get_object_or_404(ProductPrice, pk=pk)
    if request.method == "POST":
        form = ProductPriceForm(request.POST, instance=price)
        if form.is_valid():
            form.save()
            messages.success(request, "Le prix a été modifié avec succès.")
            return redirect("product_price_list")
    else:
        form = ProductPriceForm(instance=price)
    return render(request, "productprice_form.html", {"form": form})

def product_price_delete(request, pk):
    price = get_object_or_404(ProductPrice, pk=pk)
    if request.method == "POST":
        price.delete()
        messages.success(request, "Le prix a été supprimé avec succès.")
        return redirect("product_price_list")
    return render(request, "productprice_confirm_delete.html", {"product_price": price})

def product_price_list(request):
    prices = ProductPrice.objects.all()
    return render(request, "productprice_list.html", {"productprice_list": prices})

def product_price_import(request):
    """
    Importe la table ProductPrice depuis un fichier Excel
    (product, point_of_sale, value, date_from, date_to).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = row[0]
                pos_code = row[1]
                value = row[2]
                date_from = row[3]
                date_to = row[4]

                product = Product.objects.get(name=product_name)
                point_of_sale = PointOfSale.objects.get(code=pos_code)

                ProductPrice.objects.create(
                    product=product,
                    point_of_sale=point_of_sale,
                    value=value,
                    date_from=date_from,
                    date_to=date_to
                )
            messages.success(request, "Importation des ProductPrices réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("product_price_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("product_price_list")

def product_price_export(request):
    """
    Exporte la table ProductPrice dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProductPrice"

    # Ajouter l'en-tête
    fields = ["product", "point_of_sale", "value", "date_from", "date_to"]
    ws.append(fields)

    # Ajouter les données
    for price in ProductPrice.objects.all():
        row_data = [
            price.product.name,
            price.point_of_sale.code,
            price.value,
            price.date_from,
            price.date_to
        ]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="product_price_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   CRUD - CARTPRODUCTS       #
# =========================== #

def cartproducts_list(request):
    cartproducts = CartProducts.objects.all()
    return render(request, "cartproducts_list.html", {"cartproducts_list": cartproducts})

def cartproducts_add(request):
    if request.method == "POST":
        form = CartProductsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit ajouté au panier avec succès !")
            return redirect("cartproducts_list")
    else:
        form = CartProductsForm()
    return render(request, "cartproducts_form.html", {"form": form})

def cartproducts_edit(request, pk):
    cartproduct = get_object_or_404(CartProducts, pk=pk)
    if request.method == "POST":
        form = CartProductsForm(request.POST, instance=cartproduct)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit dans le panier modifié avec succès !")
            return redirect("cartproducts_list")
    else:
        form = CartProductsForm(instance=cartproduct)
    return render(request, "cartproducts_form.html", {"form": form})

def cartproducts_delete(request, pk):
    cartproduct = get_object_or_404(CartProducts, pk=pk)
    if request.method == "POST":
        cartproduct.delete()
        messages.success(request, "Produit retiré du panier avec succès !")
        return redirect("cartproducts_list")
    return render(request, "cartproducts_confirm_delete.html", {"cartproduct": cartproduct})

def cartproducts_import(request):
    """
    Importe la table CartProducts depuis un fichier Excel
    (code, name, description).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                CartProducts.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val
                )
            messages.success(request, "Importation des CartProducts réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("cartproducts_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("cartproducts_list")

def cartproducts_export(request):
    """
    Exporte la table CartProducts dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CartProducts"

    # Ajouter l'en-tête
    fields = ["product", "cart_product", "weight", "date_from", "date_to"]
    ws.append(fields)

    # Ajouter les données
    for cartproduct in CartProducts.objects.all():
        row_data = [
            cartproduct.product.name,
            cartproduct.cart_product.name,
            cartproduct.weight,
            cartproduct.date_from,
            cartproduct.date_to
        ]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cartproducts_export.xlsx"'
    wb.save(response)
    return response

# =========================== #
#   IMPORT/EXPORT EXCEL       #
# =========================== #

def export_data(request):
    """
    Exporte toutes les tables dans un seul fichier Excel, un onglet par table.
    Si un champ est un objet (ForeignKey), on convertit en str(obj) ou en id.
    """
    wb = openpyxl.Workbook()

    models = {
        "ProductType": ProductType,
        "Product": Product,
        "Wilaya": Wilaya,
        "Moughataa": Moughataa,
        "Commune": Commune,
        "PointOfSale": PointOfSale,
        "ProductPrice": ProductPrice,
        "Cart": Cart,
        "CartProducts": CartProducts,
    }

    for sheet_name, model in models.items():
        ws = wb.create_sheet(title=sheet_name)
        fields = [field.name for field in model._meta.fields]

        # Ajouter l'en-tête
        ws.append(fields)

        # Ajouter les données
        for obj in model.objects.all():
            row_data = []
            for field in fields:
                value = getattr(obj, field)

                # Convertir un objet (ForeignKey) en string ou id
                if value is not None and hasattr(value, "__class__") and "django.db.models" in str(value.__class__):
                    # Option 1 : value = value.id
                    # Option 2 : value = str(value)
                    value = str(value)

                row_data.append(value)
            ws.append(row_data)

    # Supprimer la feuille "Sheet" créée par défaut
    wb.remove(wb["Sheet"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="export_data.xlsx"'
    wb.save(response)
    return response

def import_data(request):
    """
    Importe un fichier Excel où chaque onglet correspond à une table
    (sheet_name doit être parmi ProductType, Product, etc.).
    Le code lit les champs dans l'ordre et crée l'objet.
    """
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        try:
            wb = openpyxl.load_workbook(file)
            models_dict = {
                "ProductType": ProductType,
                "Product": Product,
                "Wilaya": Wilaya,
                "Moughataa": Moughataa,
                "Commune": Commune,
                "PointOfSale": PointOfSale,
                "ProductPrice": ProductPrice,
                "Cart": Cart,
                "CartProducts": CartProducts,
            }

            for sheet_name in wb.sheetnames:
                if (sheet_name in models_dict):
                    ws = wb[sheet_name]
                    model = models_dict[sheet_name]
                    fields = [field.name for field in model._meta.fields]

                    for row in ws.iter_rows(min_row=2, values_only=True):
                        data = {}
                        for i, field_name in enumerate(fields):
                            data[field_name] = row[i]
                        model.objects.create(**data)

            messages.success(request, "Importation réussie !")
            return redirect("home")

        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")

    return render(request, "import_data.html")


# =========================== #
#   VUES D'IMPORT SPÉCIFIQUES #
# =========================== #

def cart_import(request):
    """
    Importe la table Cart depuis un fichier Excel
    (code, name, description).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                Cart.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val
                )
            messages.success(request, "Importation des Carts réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("cart_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("cart_list")

def cart_export(request):
    """
    Exporte la table Cart dans un fichier Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cart"

    # Ajouter l'en-tête
    fields = ["code", "name", "description"]
    ws.append(fields)

    # Ajouter les données
    for cart in Cart.objects.all():
        row_data = [cart.code, cart.name, cart.description]
        ws.append(row_data)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cart_export.xlsx"'
    wb.save(response)
    return response

def product_import(request):
    """
    Importe la table Product depuis un fichier Excel
    (code, name, description, unit_measure, product_type).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                unit_val = row[3]
                product_type_label = row[4]

                product_type = ProductType.objects.get(label=product_type_label)

                Product.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val,
                    unit_measure=unit_val,
                    product_type=product_type
                )
            messages.success(request, "Importation des Produits réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("product_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("product_list")

def product_type_import(request):
    """
    Importe la table ProductType depuis un fichier Excel
    (code, name, description).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                desc_val = row[2]
                ProductType.objects.create(
                    code=code_val,
                    name=name_val,
                    description=desc_val
                )
            messages.success(request, "Importation des ProductTypes réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("product_type_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("product_type_list")

def moughataa_import(request):
    """
    Importe la table Moughataa depuis un fichier Excel
    (code, label, wilaya).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                label_val = row[1]
                wilaya_name = row[2]

                wilaya = Wilaya.objects.get(name=wilaya_name)

                Moughataa.objects.create(
                    code=code_val,
                    label=label_val,
                    wilaya=wilaya
                )
            messages.success(request, "Importation des Moughataas réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("moughataa_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("moughataa_list")

def commune_import(request):
    """
    Importe la table Commune depuis un fichier Excel
    (code, name, moughataa).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                code_val = row[0]
                name_val = row[1]
                moughataa_name = row[2]

                moughataa = Moughataa.objects.get(name=moughataa_name)

                Commune.objects.create(
                    code=code_val,
                    name=name_val,
                    moughataa=moughataa
                )
            messages.success(request, "Importation des Communes réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("commune_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("commune_list")

def product_price_import(request):
    """
    Importe la table ProductPrice depuis un fichier Excel
    (product, point_of_sale, value, date_from, date_to).
    """
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active  
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = row[0]
                pos_code = row[1]
                value = row[2]
                date_from = row[3]
                date_to = row[4]

                product = Product.objects.get(name=product_name)
                point_of_sale = PointOfSale.objects.get(code=pos_code)

                ProductPrice.objects.create(
                    product=product,
                    point_of_sale=point_of_sale,
                    value=value,
                    date_from=date_from,
                    date_to=date_to
                )
            messages.success(request, "Importation des ProductPrices réussie !")
        except Exception as e:
            messages.error(request, f"Erreur lors de l'importation : {e}")
        return redirect("product_price_list")

    messages.warning(request, "Aucun fichier détecté.")
    return redirect("product_price_list")

# ... Etc. pour wilaya_import, moughataa_import, commune_import, point_of_sale_import,
# product_price_import, cartproducts_import (voir code précédent) ...

